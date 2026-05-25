"""Build Home Mortgage Calculator workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.comments import Comment

# Styles
BLUE = Font(name="Arial", color="0000FF", size=11)
BLUE_BOLD = Font(name="Arial", color="0000FF", size=11, bold=True)
BLACK = Font(name="Arial", color="000000", size=11)
BLACK_BOLD = Font(name="Arial", color="000000", size=11, bold=True)
GREEN = Font(name="Arial", color="008000", size=11)
GREEN_BOLD = Font(name="Arial", color="008000", size=11, bold=True)
HEADER = Font(name="Arial", color="FFFFFF", size=11, bold=True)
TITLE = Font(name="Arial", color="000000", size=16, bold=True)
SUBTITLE = Font(name="Arial", color="404040", size=11, italic=True)
SECTION = Font(name="Arial", color="000000", size=13, bold=True)

YELLOW_FILL = PatternFill("solid", start_color="FFFF00")
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
BAND_FILL = PatternFill("solid", start_color="F2F2F2")
LIGHT_GREEN = PatternFill("solid", start_color="E2EFDA")
LIGHT_BLUE = PatternFill("solid", start_color="DDEBF7")

CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.00%;(0.00%);"-"'
INT_FMT = '#,##0;(#,##0);"-"'

THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cell(cell, font=None, fill=None, fmt=None, align=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    if border: cell.border = border


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # ==================== INPUTS SHEET ====================
    ws = wb.active
    ws.title = "Inputs"
    set_widths(ws, [38, 18, 30])

    ws["A1"] = "Home Mortgage Calculator — Inputs"
    ws["A1"].font = TITLE
    ws["A2"] = "Blue cells = editable assumptions. All other sheets pull from this one."
    ws["A2"].font = SUBTITLE

    row = 4
    ws.cell(row=row, column=1, value="LOAN").font = SECTION
    row += 1

    inputs = [
        ("Home price ($)",           "home_price",    500_000, CURRENCY_FMT, "Purchase price of the home"),
        ("Down payment (%)",         "down_pct",      0.20,    PCT_FMT, "% of price paid up-front"),
        ("Interest rate (annual %)", "rate",          0.065,   PCT_FMT, "APR — fixed-rate mortgage"),
        ("Loan term (years)",        "term",          30,      "0", "Common: 15, 20, 30"),
        ("RECURRING COSTS", None, None, None, None),
        ("Property tax (annual %)",  "tax_pct",       0.012,   PCT_FMT, "Typical: 0.5–2.0% of home value"),
        ("Home insurance ($/year)",  "insurance_yr",  1_800,   CURRENCY_FMT, "Annual premium"),
        ("HOA ($/month)",            "hoa_mo",        0,       CURRENCY_FMT, "$0 if no HOA"),
        ("PMI rate (annual %)",      "pmi_pct",       0.006,   PCT_FMT, "Usually 0.3–1.5%; only applies if down < 20%"),
        ("APPRECIATION & COSTS", None, None, None, None),
        ("Home appreciation (annual %)", "appr",      0.035,   PCT_FMT, "Long-run US avg ~3–4%"),
        ("Maintenance (% of home value/yr)", "maint", 0.010,   PCT_FMT, "Rule of thumb: 1%/yr"),
        ("Closing costs (% of price)", "close_pct",   0.030,   PCT_FMT, "Typical: 2–5%"),
        ("Selling costs (% of sale)",  "sell_pct",    0.060,   PCT_FMT, "Typical: 6% (agent fees etc)"),
        ("EXTRA PAYMENTS", None, None, None, None),
        ("Extra monthly principal ($)", "extra",      0,       CURRENCY_FMT, "Additional principal each month"),
        ("RENT VS. BUY", None, None, None, None),
        ("Current rent ($/month)",   "rent_mo",       2_500,   CURRENCY_FMT, "What you'd pay if renting"),
        ("Annual rent increase (%)", "rent_inc",      0.030,   PCT_FMT, "Typical: 2–4%"),
        ("Investment return (%)",    "inv_ret",       0.070,   PCT_FMT, "Opportunity cost of down payment"),
        ("AFFORDABILITY", None, None, None, None),
        ("Other monthly debts ($)",  "debts_mo",      0,       CURRENCY_FMT, "Car, student loans, credit card min etc."),
        ("Custom front-end DTI (%)", "dti_custom",    0.28,    PCT_FMT, "Your housing-to-income ceiling"),
        ("ESCALATION", None, None, None, None),
        ("Property tax growth (annual %)", "tax_growth",  0.020,  PCT_FMT, "How fast tax grows; CA Prop 13 caps at 2%"),
        ("Insurance growth (annual %)",    "ins_growth",  0.050,  PCT_FMT, "Often 4-7%; higher in climate-affected areas"),
        ("CASH TO CLOSE", None, None, None, None),
        ("Initial repairs / move-in ($)",  "movein",     5_000,   CURRENCY_FMT, "Lump-sum at move-in: furniture, repairs etc."),
        ("Required reserves (months PITI)", "reserve_mos", 2,     "0", "Lenders want 2-6 months of PITI in reserves"),
        ("SAVINGS & ASSETS", None, None, None, None),
        ("Current liquid assets ($)",      "liquid",     150_000, CURRENCY_FMT, "Cash + brokerage + savings (excl. retirement)"),
        ("Monthly savings ($/mo)",         "save_mo",    2_000,   CURRENCY_FMT, "What you can save toward the down payment"),
        ("Monthly household expenses ($)", "expenses_mo", 6_000,  CURRENCY_FMT, "Food, utilities, transport — excl. mortgage"),
        ("Emergency fund target (months)", "emerg_mos",  6,       "0", "Recommended: 6 months of expenses"),
        ("Max % of liquid for down payment", "max_liq_pct", 0.50,  PCT_FMT, "Concentration cap — don't lock more than this share into one asset"),
        ("TAXES", None, None, None, None),
        ("Filing status (M=married, S=single)", "filing", "M",   "@", "M = MFJ ($500K cap gains excl.); S = Single ($250K)"),
        ("Marginal federal tax rate (%)",  "marg_rate",  0.24,    PCT_FMT, "Your top federal bracket"),
        ("Other SALT taxes paid ($/yr)",   "salt_other", 8_000,   CURRENCY_FMT, "Income + other state/local (excl. property tax)"),
        ("LTCG rate at sale (%)",          "ltcg_rate",  0.15,    PCT_FMT, "Long-term cap gains rate: 0/15/20% federal"),
        ("RETIREMENT", None, None, None, None),
        ("Gross annual income ($)",        "income",     150_000,  CURRENCY_FMT, "Household gross income"),
        ("Current 401k contribution (%)",  "retire_pct", 0.10,     PCT_FMT, "Your pre-tax 401k contribution"),
        ("Employer match (%)",             "emp_match",  0.50,     PCT_FMT, "Employer matches this % of your contribution"),
        ("Match cap (% of salary)",        "match_cap",  0.06,     PCT_FMT, "Employer only matches up to this % of salary"),
        ("Expected investment return (%)", "retire_ret", 0.07,     PCT_FMT, "Long-term expected return"),
        ("Years to retirement",            "retire_yrs", 25,       "0", "Years until retirement"),
        ("REFINANCE", None, None, None, None),
        ("Include refi? (1=yes, 0=no)",     "refi_enabled", 0,       "0", "Toggle refi simulation on/off"),
        ("Refinance at year",               "refi_year",   5,       "0", "When you'd refi (1-30)"),
        ("New interest rate (%)",           "refi_rate",   0.055,   PCT_FMT, "Rate on new loan"),
        ("New term (years)",                "refi_term",   30,      "0", "30 resets clock; 15/20 keeps shorter trajectory"),
        ("Refi closing costs (% of new loan)", "refi_close", 0.020, PCT_FMT, "Typical 2-5%"),
        ("Cash-out amount ($)",             "cashout",     0,       CURRENCY_FMT, "Equity converted to cash"),
        ("Roll closing into loan (1=yes,0=no)", "roll_closing", 0,  "0", "If 1, closing added to loan balance"),
    ]

    # Track named cells
    named = {}

    for label, name, val, fmt, note in inputs:
        if name is None:
            ws.cell(row=row, column=1, value=label).font = SECTION
            row += 1
            continue
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = BLACK
        c = ws.cell(row=row, column=2, value=val)
        c.font = BLUE
        c.fill = YELLOW_FILL
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        c.border = BOX
        n = ws.cell(row=row, column=3, value=note)
        n.font = SUBTITLE
        named[name] = f"Inputs!$B${row}"
        row += 1

    # Derived constants on Inputs sheet
    row += 1
    ws.cell(row=row, column=1, value="DERIVED").font = SECTION
    row += 1
    derived_start = row

    # Down payment $
    ws.cell(row=row, column=1, value="Down payment ($)").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['home_price']}*{named['down_pct']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['down_amt'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Loan amount ($)").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['home_price']}-{named['down_amt']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['loan'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly rate").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['rate']}/12")
    c.font = BLACK; c.number_format = "0.00000%"; c.border = BOX
    named['monthly_rate'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Number of payments").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['term']}*12")
    c.font = BLACK; c.number_format = "0"; c.border = BOX
    named['n_pmts'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly P&I").font = BLACK_BOLD
    c = ws.cell(row=row, column=2,
                value=f"=-PMT({named['monthly_rate']},{named['n_pmts']},{named['loan']})")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.fill = LIGHT_GREEN; c.border = BOX
    named['pi'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly property tax").font = BLACK
    c = ws.cell(row=row, column=2,
                value=f"={named['tax_pct']}*{named['home_price']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['tax_mo'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly insurance").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['insurance_yr']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['ins_mo'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly PMI (initial)").font = BLACK
    c = ws.cell(row=row, column=2,
                value=f"={named['pmi_pct']}*{named['loan']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['pmi_mo'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Monthly HOA").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['hoa_mo']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['hoa_mo_d'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="Closing costs ($)").font = BLACK
    c = ws.cell(row=row, column=2,
                value=f"={named['home_price']}*{named['close_pct']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['closing'] = f"Inputs!$B${row}"; row += 1

    ws.cell(row=row, column=1, value="PMI cutoff (80% LTV equity)").font = BLACK
    c = ws.cell(row=row, column=2, value=f"={named['home_price']}*0.20")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['pmi_cutoff'] = f"Inputs!$B${row}"; row += 1

    # Summary section
    row += 2
    ws.cell(row=row, column=1, value="MONTHLY PAYMENT SUMMARY (Year 1)").font = SECTION
    row += 1

    summary_items = [
        ("Principal & Interest", named['pi']),
        ("Property tax",         named['tax_mo']),
        ("Insurance",            named['ins_mo']),
        ("PMI (if down < 20%)",  f"IF({named['down_pct']}<0.2,{named['pmi_mo']},0)"),
        ("HOA",                  named['hoa_mo_d']),
    ]
    sum_start = row
    for label, ref in summary_items:
        ws.cell(row=row, column=1, value=label).font = BLACK
        if "IF" in ref:
            c = ws.cell(row=row, column=2, value=f"={ref}")
        else:
            c = ws.cell(row=row, column=2, value=f"={ref}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
        row += 1
    ws.cell(row=row, column=1, value="Total monthly").font = BLACK_BOLD
    c = ws.cell(row=row, column=2, value=f"=SUM(B{sum_start}:B{row-1})")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.fill = LIGHT_GREEN; c.border = BOX
    named['monthly_total'] = f"Inputs!$B${row}"
    row += 1

    # ==================== AMORTIZATION SHEET ====================
    ws_a = wb.create_sheet("Amortization")
    set_widths(ws_a, [10, 16, 16, 16, 16, 16, 16, 16, 16])

    ws_a["A1"] = "Amortization Schedule"
    ws_a["A1"].font = TITLE
    ws_a["A2"] = "Formulas reference the Inputs sheet. Edit assumptions there to update everything."
    ws_a["A2"].font = SUBTITLE

    headers = ["Month", "Year", "Beginning Balance", "Payment", "Principal", "Interest",
               "Extra Principal", "PMI Active?", "Ending Balance"]
    header_row(ws_a, 4, headers)

    n_pmts_max = 360  # support up to 30 yrs
    start_row = 5
    for i in range(n_pmts_max):
        r = start_row + i
        m = i + 1
        ws_a.cell(row=r, column=1, value=m)
        ws_a.cell(row=r, column=2, value=f"=CEILING(A{r}/12,1)").number_format = "0"
        # Beginning balance
        if i == 0:
            ws_a.cell(row=r, column=3, value=f"={named['loan']}")
        else:
            ws_a.cell(row=r, column=3, value=f"=I{r-1}")
        # Payment (only if balance > 0 and within term)
        ws_a.cell(row=r, column=4,
                  value=f"=IF(AND(C{r}>0.01,A{r}<={named['n_pmts']}),{named['pi']},0)")
        # Interest
        ws_a.cell(row=r, column=6,
                  value=f"=IF(C{r}>0.01,C{r}*{named['monthly_rate']},0)")
        # Principal (capped by balance)
        ws_a.cell(row=r, column=5,
                  value=f"=IF(D{r}=0,0,MIN(D{r}-F{r},C{r}))")
        # Extra principal (capped by remaining)
        ws_a.cell(row=r, column=7,
                  value=f"=IF(C{r}-E{r}>0,MIN({named['extra']},C{r}-E{r}),0)")
        # PMI active check: based on cumulative principal paid
        # PMI is active if (down + cumulative principal) < 20% of home price
        ws_a.cell(row=r, column=8,
                  value=f"=IF({named['down_amt']}+SUM(E$5:E{r})+SUM(G$5:G{r-1 if i>0 else r})<{named['pmi_cutoff']},1,0)")
        # Ending balance
        ws_a.cell(row=r, column=9, value=f"=MAX(0,C{r}-E{r}-G{r})")

        # Format
        for col in [3,4,5,6,7,9]:
            ws_a.cell(row=r, column=col).number_format = CURRENCY_FMT
        ws_a.cell(row=r, column=8).number_format = '"Yes";;"No"'
        ws_a.cell(row=r, column=8).alignment = Alignment(horizontal="center")

    # Freeze header
    ws_a.freeze_panes = "A5"

    # Totals row at bottom
    tot_row = start_row + n_pmts_max + 1
    ws_a.cell(row=tot_row, column=1, value="TOTAL").font = BLACK_BOLD
    for col, lbl in [(4, "payments"), (5, "principal"), (6, "interest"), (7, "extra")]:
        c = ws_a.cell(row=tot_row, column=col,
                      value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{tot_row-1})")
        c.font = BLACK_BOLD
        c.number_format = CURRENCY_FMT
        c.fill = LIGHT_GREEN
    named['total_interest'] = f"Amortization!$F${tot_row}"
    named['total_principal'] = f"Amortization!$E${tot_row}"
    named['total_extra'] = f"Amortization!$G${tot_row}"
    named['total_payments'] = f"Amortization!$D${tot_row}"

    # ==================== APPRECIATION SHEET ====================
    ws_p = wb.create_sheet("Appreciation")
    set_widths(ws_p, [8, 16, 16, 16, 16, 16, 16, 16])
    ws_p["A1"] = "Home Appreciation & Equity"
    ws_p["A1"].font = TITLE
    ws_p["A2"] = "Year-by-year asset projection. Net equity = appreciated value − loan balance − selling costs."
    ws_p["A2"].font = SUBTITLE

    headers = ["Year", "Home Value", "Loan Balance", "Gross Equity",
               "Selling Costs", "Net Equity", "Cum. Interest", "Cum. Principal"]
    header_row(ws_p, 4, headers)

    for y in range(1, 31):
        r = 4 + y
        ws_p.cell(row=r, column=1, value=y)
        # Home value = price * (1+appr)^y
        ws_p.cell(row=r, column=2,
                  value=f"={named['home_price']}*(1+{named['appr']})^A{r}")
        # Loan balance: pull from amortization sheet at month=y*12
        # Use INDEX into amortization column I
        ws_p.cell(row=r, column=3,
                  value=f"=IFERROR(INDEX(Amortization!I:I,MIN({4+y*12},{4+n_pmts_max})),0)")
        # Gross equity
        ws_p.cell(row=r, column=4, value=f"=B{r}-C{r}")
        # Selling costs
        ws_p.cell(row=r, column=5, value=f"=B{r}*{named['sell_pct']}")
        # Net equity
        ws_p.cell(row=r, column=6, value=f"=D{r}-E{r}")
        # Cumulative interest up through month y*12
        ws_p.cell(row=r, column=7,
                  value=f"=SUM(Amortization!F$5:INDEX(Amortization!F:F,MIN({4+y*12},{4+n_pmts_max})))")
        # Cumulative principal
        ws_p.cell(row=r, column=8,
                  value=f"=SUM(Amortization!E$5:INDEX(Amortization!E:E,MIN({4+y*12},{4+n_pmts_max})))+SUM(Amortization!G$5:INDEX(Amortization!G:G,MIN({4+y*12},{4+n_pmts_max})))")

        for col in [2,3,4,5,6,7,8]:
            ws_p.cell(row=r, column=col).number_format = CURRENCY_FMT
        if y % 5 == 0:
            for col in range(1, 9):
                ws_p.cell(row=r, column=col).fill = LIGHT_BLUE

    # Add a line chart
    chart = LineChart()
    chart.title = "Home Value vs. Loan Balance vs. Net Equity"
    chart.y_axis.title = "Dollars"
    chart.x_axis.title = "Year"
    data = Reference(ws_p, min_col=2, max_col=6, min_row=4, max_row=34)
    cats = Reference(ws_p, min_col=1, min_row=5, max_row=34)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws_p.add_chart(chart, "J4")

    # ==================== RENT VS BUY SHEET ====================
    ws_r = wb.create_sheet("Rent vs Buy")
    set_widths(ws_r, [8, 16, 16, 16, 16, 16, 16, 16, 16])
    ws_r["A1"] = "Rent vs. Buy Comparison"
    ws_r["A1"].font = TITLE
    ws_r["A2"] = "Compares net wealth path of buying (net equity − ownership costs) vs. renting (down payment invested − rent paid)."
    ws_r["A2"].font = SUBTITLE

    headers = ["Year", "Annual Buy Cost", "Cum. Buy Cost", "Annual Rent",
               "Cum. Rent", "Investment Acct", "Buy Net Wealth", "Rent Net Wealth", "Buy − Rent"]
    header_row(ws_r, 4, headers)

    for y in range(1, 31):
        r = 4 + y
        ws_r.cell(row=r, column=1, value=y)
        # Annual buy cost = sum of monthly P&I + tax + ins + pmi + hoa + maintenance for that year
        # Approx: use total monthly summary × 12 + maintenance% × current home value
        # Actually we want: 12 × (P&I + tax + ins + PMI(if active) + HOA) + maint × home value
        # PMI active if cum_principal+down < pmi_cutoff at year y-1
        # Simpler: pull P&I from amortization total for that year window
        m_start = 4 + (y-1)*12 + 1
        m_end = 4 + y*12
        ws_r.cell(row=r, column=2,
                  value=f"=SUM(Amortization!D{m_start}:D{m_end})+SUM(Amortization!G{m_start}:G{m_end})"
                        f"+12*({named['tax_mo']}*(1+{named['tax_growth']})^(A{r}-1)"
                        f"+{named['ins_mo']}*(1+{named['ins_growth']})^(A{r}-1)"
                        f"+{named['hoa_mo_d']})"
                        f"+SUM(Amortization!H{m_start}:H{m_end})*{named['pmi_mo']}"
                        f"+{named['maint']}*Appreciation!B{r}")
        # Cumulative buy cost
        if y == 1:
            ws_r.cell(row=r, column=3, value=f"={named['closing']}+B{r}")
        else:
            ws_r.cell(row=r, column=3, value=f"=C{r-1}+B{r}")

        # Annual rent
        ws_r.cell(row=r, column=4,
                  value=f"={named['rent_mo']}*12*(1+{named['rent_inc']})^(A{r}-1)")
        # Cumulative rent
        if y == 1:
            ws_r.cell(row=r, column=5, value=f"=D{r}")
        else:
            ws_r.cell(row=r, column=5, value=f"=E{r-1}+D{r}")

        # Investment account: (down + closing) grown at investment return
        ws_r.cell(row=r, column=6,
                  value=f"=({named['down_amt']}+{named['closing']})*(1+{named['inv_ret']})^A{r}")

        # Buy net wealth = Net equity (from appreciation) − cumulative buy cost (excluding dp because dp is in equity)
        # Buy gives up cash for dp+closing+payments; gets net equity.
        # Buy net = NetEquity - (CumBuyCost + DownPayment - DownPayment back via equity)
        # Cleaner: Buy net = NetEquity - CumBuyCost (since CumBuyCost includes closing; dp is recovered through equity)
        # Net financial position relative to starting cash (down+closing both spent):
        # Buy = NetEquity_y - (CumBuyCost - down_payment_recovered_in_equity) ...
        # Simpler approach: define Buy Net = NetEquity - CumOperatingCosts - closing
        # where CumOperatingCosts excludes principal paydown (already in equity).
        # Cleanest: track wealth.
        # Starting wealth: 0
        # Buy path: spent (dp+closing) -> have (homeValue@0 with loan). Each yr pay P&I+tax+ins+pmi+hoa+maint, value appreciates, principal reduces.
        # Net wealth (buy) = NetEquity - cumulative cash paid (P&I + tax + ins + pmi + hoa + maint + closing)
        # Note: dp itself moved into equity so don't subtract it again.
        ws_r.cell(row=r, column=7,
                  value=f"=Appreciation!F{r}-C{r}")
        # Rent net wealth = InvestmentAcct - CumulativeRent
        ws_r.cell(row=r, column=8, value=f"=F{r}-E{r}")
        # Diff
        ws_r.cell(row=r, column=9, value=f"=G{r}-H{r}")

        for col in [2,3,4,5,6,7,8,9]:
            ws_r.cell(row=r, column=col).number_format = CURRENCY_FMT

    # Chart
    chart = LineChart()
    chart.title = "Rent vs. Buy — Net Wealth Over Time"
    chart.y_axis.title = "Dollars"
    chart.x_axis.title = "Year"
    data = Reference(ws_r, min_col=7, max_col=9, min_row=4, max_row=34)
    cats = Reference(ws_r, min_col=1, min_row=5, max_row=34)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws_r.add_chart(chart, "K4")

    # ==================== AFFORDABILITY SHEET ====================
    ws_af = wb.create_sheet("Affordability")
    set_widths(ws_af, [34, 16, 16, 18, 18, 40])
    ws_af["A1"] = "Affordability — Minimum Required Household Income"
    ws_af["A1"].font = TITLE
    ws_af["A2"] = "Back-solves the gross annual income needed for this payment under various DTI standards."
    ws_af["A2"].font = SUBTITLE

    # Quick inputs / monthly basis
    row = 4
    ws_af.cell(row=row, column=1, value="Monthly housing (PITI + PMI + HOA)").font = BLACK_BOLD
    c = ws_af.cell(row=row, column=2, value=f"={named['monthly_total']}")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.fill = LIGHT_GREEN; c.border = BOX
    named['af_housing'] = f"Affordability!$B${row}"
    row += 1
    ws_af.cell(row=row, column=1, value="Other monthly debts").font = BLACK
    c = ws_af.cell(row=row, column=2, value=f"={named['debts_mo']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['af_debts'] = f"Affordability!$B${row}"
    row += 1
    ws_af.cell(row=row, column=1, value="Total monthly obligations").font = BLACK_BOLD
    c = ws_af.cell(row=row, column=2, value=f"={named['af_housing']}+{named['af_debts']}")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.fill = LIGHT_BLUE; c.border = BOX
    named['af_total'] = f"Affordability!$B${row}"
    row += 2

    # Standards table
    header_row(ws_af, row, ["Standard", "Front-end %", "Back-end %", "Req. gross/year", "Req. gross/month", "Comment"])
    row += 1
    standards = [
        ("Conservative (Ramsey 25%)", 0.25, 0.25, "Most conservative; leaves room for everything else."),
        ("Classic 28/36 rule",         0.28, 0.36, "Conventional lender baseline."),
        ("FHA upper limit (31/43)",    0.31, 0.43, "Stretched but qualifies for FHA."),
        ("Aggressive (35/45)",         0.35, 0.45, "Risky — leaves little margin for shocks."),
    ]
    for name_s, fe, be, comment in standards:
        ws_af.cell(row=row, column=1, value=name_s).font = BLACK
        ws_af.cell(row=row, column=2, value=fe).number_format = PCT_FMT
        ws_af.cell(row=row, column=3, value=be).number_format = PCT_FMT
        # Required monthly income = MAX(housing/front, total/back)
        ws_af.cell(row=row, column=4,
                   value=f"=MAX({named['af_housing']}/B{row},{named['af_total']}/C{row})*12")
        ws_af.cell(row=row, column=5,
                   value=f"=MAX({named['af_housing']}/B{row},{named['af_total']}/C{row})")
        ws_af.cell(row=row, column=4).number_format = CURRENCY_FMT
        ws_af.cell(row=row, column=5).number_format = CURRENCY_FMT
        ws_af.cell(row=row, column=6, value=comment).font = SUBTITLE
        for col in range(1, 7):
            ws_af.cell(row=row, column=col).border = BOX
        row += 1

    # Custom row — uses user's chosen DTI
    ws_af.cell(row=row, column=1, value="Custom (from Inputs)").font = BLACK_BOLD
    ws_af.cell(row=row, column=2, value=f"={named['dti_custom']}").number_format = PCT_FMT
    ws_af.cell(row=row, column=3, value=f"={named['dti_custom']}+0.08").number_format = PCT_FMT
    ws_af.cell(row=row, column=4,
               value=f"=MAX({named['af_housing']}/B{row},{named['af_total']}/C{row})*12")
    ws_af.cell(row=row, column=5,
               value=f"=MAX({named['af_housing']}/B{row},{named['af_total']}/C{row})")
    ws_af.cell(row=row, column=4).number_format = CURRENCY_FMT
    ws_af.cell(row=row, column=5).number_format = CURRENCY_FMT
    ws_af.cell(row=row, column=4).fill = LIGHT_GREEN
    ws_af.cell(row=row, column=5).fill = LIGHT_GREEN
    ws_af.cell(row=row, column=6, value="Your custom front-end. Back-end = front + 8 pts.").font = SUBTITLE
    for col in range(1, 7):
        ws_af.cell(row=row, column=col).border = BOX
    row += 2

    ws_af.cell(row=row, column=1, value="Formula: required_income = MAX( housing/front_DTI, total_obligations/back_DTI ) × 12").font = SUBTITLE

    # ==================== CASH TO CLOSE SHEET ====================
    ws_cc = wb.create_sheet("Cash to Close")
    set_widths(ws_cc, [40, 18, 18, 40])
    ws_cc["A1"] = "Cash to Close"
    ws_cc["A1"].font = TITLE
    ws_cc["A2"] = "Cash needed at closing and recommended liquid retention."
    ws_cc["A2"].font = SUBTITLE

    r = 4
    ws_cc.cell(row=r, column=1, value="STEP 1 — Cash needed at closing").font = SECTION
    r += 1

    items = [
        ("Down payment",          f"={named['down_amt']}",         "Spent at close"),
        ("Closing costs",         f"={named['closing']}",          "Spent at close"),
        ("Move-in / repairs",     f"={named['movein']}",            "Spent at close"),
        ("Lender reserves",       f"={named['monthly_total']}*{named['reserve_mos']}",
                                                                    "Stays in account post-close"),
    ]
    for label, formula, note in items:
        ws_cc.cell(row=r, column=1, value=label).font = BLACK
        c = ws_cc.cell(row=r, column=2, value=formula)
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
        ws_cc.cell(row=r, column=4, value=note).font = SUBTITLE
        r += 1

    ws_cc.cell(row=r, column=1, value="Total cash needed at close").font = BLACK_BOLD
    c = ws_cc.cell(row=r, column=2, value=f"=SUM(B5:B{r-1})")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX
    named['cash_close'] = f"'Cash to Close'!$B${r}"
    r += 2

    # Liquid after close: reserves stay in account, so we don't subtract them
    ws_cc.cell(row=r, column=1, value="Liquid remaining after close").font = BLACK
    c = ws_cc.cell(row=r, column=2,
                   value=f"={named['liquid']}-{named['down_amt']}-{named['closing']}-{named['movein']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_cc.cell(row=r, column=4, value="Includes lender reserves; those stay liquid").font = SUBTITLE
    named['liquid_after'] = f"'Cash to Close'!$B${r}"
    r += 2

    # STEP 2 — Recommended retention
    ws_cc.cell(row=r, column=1, value="STEP 2 — Recommended liquid retention").font = SECTION
    r += 1
    ws_cc.cell(row=r, column=1, value="Monthly cost of living (housing + other)").font = BLACK
    c = ws_cc.cell(row=r, column=2,
                   value=f"={named['expenses_mo']}+{named['monthly_total']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['monthly_outflows'] = f"'Cash to Close'!$B${r}"
    r += 1
    ws_cc.cell(row=r, column=1, value="Emergency fund target").font = BLACK
    c = ws_cc.cell(row=r, column=2,
                   value=f"={named['monthly_outflows']}*{named['emerg_mos']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_cc.cell(row=r, column=4,
               value=f"Full living costs × emergency months").font = SUBTITLE
    named['emerg_target'] = f"'Cash to Close'!$B${r}"
    r += 1
    ws_cc.cell(row=r, column=1, value="Lender reserves (subset of emergency fund)").font = BLACK
    c = ws_cc.cell(row=r, column=2,
                   value=f"={named['monthly_total']}*{named['reserve_mos']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1
    ws_cc.cell(row=r, column=1, value="Total liquid to retain").font = BLACK_BOLD
    c = ws_cc.cell(row=r, column=2,
                   value=f"=MAX({named['emerg_target']},{named['monthly_total']}*{named['reserve_mos']})")
    c.font = BLACK_BOLD; c.fill = LIGHT_BLUE; c.number_format = CURRENCY_FMT; c.border = BOX
    named['retain_floor'] = f"'Cash to Close'!$B${r}"

    # ==================== BUY TODAY SHEET ====================
    ws_sv = wb.create_sheet("Buy Today")
    set_widths(ws_sv, [38, 18, 18, 40])
    ws_sv["A1"] = "Buy Today Analysis"
    ws_sv["A1"].font = TITLE
    ws_sv["A2"] = "What your position looks like if you buy today with current assets."
    ws_sv["A2"].font = SUBTITLE

    # Liquidity check
    r = 4
    ws_sv.cell(row=r, column=1, value="LIQUIDITY CHECK").font = SECTION
    r += 1
    ws_sv.cell(row=r, column=1, value="Current liquid assets").font = BLACK
    c = ws_sv.cell(row=r, column=2, value=f"={named['liquid']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1
    ws_sv.cell(row=r, column=1, value="% of liquid used for closing").font = BLACK
    c = ws_sv.cell(row=r, column=2,
                   value=f"=IF({named['liquid']}>0,{named['cash_close']}/{named['liquid']},1)")
    c.font = BLACK; c.number_format = PCT_FMT; c.border = BOX
    r += 1
    ws_sv.cell(row=r, column=1, value="Liquid remaining after close").font = BLACK_BOLD
    c = ws_sv.cell(row=r, column=2, value=f"={named['liquid_after']}")
    c.font = BLACK_BOLD; c.fill = LIGHT_BLUE; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1
    ws_sv.cell(row=r, column=1, value="Emergency fund target").font = BLACK
    c = ws_sv.cell(row=r, column=2, value=f"={named['emerg_target']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 2

    # Concentration cap
    ws_sv.cell(row=r, column=1, value="CONCENTRATION CAP").font = SECTION
    r += 1
    ws_sv.cell(row=r, column=1, value="Planned down payment").font = BLACK
    c = ws_sv.cell(row=r, column=2, value=f"={named['down_amt']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1
    ws_sv.cell(row=r, column=1, value="Liquid required at purchase").font = BLACK_BOLD
    c = ws_sv.cell(row=r, column=2,
                   value=f"=IF({named['max_liq_pct']}>0,{named['down_amt']}/{named['max_liq_pct']},0)")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_sv.cell(row=r, column=3, value="Planned DP ÷ Max % of liquid").font = SUBTITLE
    named['req_liquid_for_dp'] = f"'Buy Today'!$B${r}"
    r += 1
    ws_sv.cell(row=r, column=1, value="Current liquid").font = BLACK
    c = ws_sv.cell(row=r, column=2, value=f"={named['liquid']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1
    ws_sv.cell(row=r, column=1, value="Surplus / shortfall").font = BLACK_BOLD
    c = ws_sv.cell(row=r, column=2,
                   value=f"={named['liquid']}-{named['req_liquid_for_dp']}")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_sv.cell(row=r, column=3, value="Positive = surplus. Negative = need more liquid.").font = SUBTITLE
    named['liquid_gap'] = f"'Buy Today'!$B${r}"
    r += 1
    # Time to reach goal
    ws_sv.cell(row=r, column=1, value="Monthly return (for savings calc)").font = BLACK
    c = ws_sv.cell(row=r, column=2, value=f"=(1+{named['inv_ret']})^(1/12)-1")
    c.font = BLACK; c.number_format = '0.0000%'; c.border = BOX
    named['monthly_ret'] = f"'Buy Today'!$B${r}"
    r += 1
    ws_sv.cell(row=r, column=1, value="Time to reach goal").font = BLACK_BOLD
    c = ws_sv.cell(row=r, column=2,
                   value=f'=IF({named["liquid_gap"]}>=0,"Already there",IF({named["save_mo"]}<=0,"Need savings",IFERROR(NPER({named["monthly_ret"]},-{named["save_mo"]},0,-{named["liquid_gap"]})/12,0)&" yr"))')
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.border = BOX
    ws_sv.cell(row=r, column=3, value="Years to save to reach required liquid").font = SUBTITLE
    r += 1
    ws_sv.cell(row=r, column=1, value="Status").font = BLACK_BOLD
    c = ws_sv.cell(row=r, column=2,
                   value=f'=IF({named["liquid"]}>={named["req_liquid_for_dp"]},"OK — within concentration cap","Below — save more, lower DP, or raise cap")')
    c.font = BLACK_BOLD; c.alignment = Alignment(horizontal="center"); c.border = BOX

    # ==================== TAX BENEFITS SHEET ====================
    ws_tx = wb.create_sheet("Tax Benefits")
    set_widths(ws_tx, [38, 18, 18, 18, 40])
    ws_tx["A1"] = "Tax Benefits — Mortgage Interest, SALT, and Capital Gains"
    ws_tx["A1"].font = TITLE
    ws_tx["A2"] = "Federal benefits from itemizing vs standard, plus $250K/$500K Section 121 exclusion on sale."
    ws_tx["A2"].font = SUBTITLE

    r = 4
    # Year-by-year deduction table for years 1, 5, 10, 20, 30
    header_row(ws_tx, r, ["Item", "Year 1", "Year 5", "Year 10", "Year 30"])
    r += 1

    # Mortgage interest in each year (sum of 12 months from amortization)
    def yr_int_formula(y):
        m_start = 4 + (y-1)*12 + 1
        m_end = 4 + y*12
        return f"=SUM(Amortization!F{m_start}:F{m_end})"

    yrs = [1, 5, 10, 30]
    ws_tx.cell(row=r, column=1, value="Mortgage interest paid").font = BLACK
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i, value=yr_int_formula(y))
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_tx.cell(row=r, column=6, value="From amortization").font = SUBTITLE
    int_row = r
    r += 1

    # Deductible interest (capped at $750K loan ratio)
    ws_tx.cell(row=r, column=1, value="Deductible interest (post-TCJA $750K cap)").font = BLACK
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f"={get_column_letter(2+i)}{int_row}*IF({named['loan']}>750000,750000/{named['loan']},1)")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_tx.cell(row=r, column=6, value="If loan > $750K, scale interest proportionally").font = SUBTITLE
    int_ded_row = r
    r += 1

    # Property tax in each year (escalated)
    ws_tx.cell(row=r, column=1, value="Property tax (escalated)").font = BLACK
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f"={named['tax_pct']}*{named['home_price']}*(1+{named['tax_growth']})^({y}-1)")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_tx.cell(row=r, column=6, value="Grows at tax-growth rate from Inputs").font = SUBTITLE
    pt_row = r
    r += 1

    # SALT deductible
    ws_tx.cell(row=r, column=1, value="SALT deductible (capped at $10K)").font = BLACK
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f"=MIN(10000,{get_column_letter(2+i)}{pt_row}+{named['salt_other']})")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_tx.cell(row=r, column=6, value="Property tax + other SALT, capped at $10K").font = SUBTITLE
    salt_row = r
    r += 1

    # Itemized total
    ws_tx.cell(row=r, column=1, value="Itemized total").font = BLACK_BOLD
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f"={get_column_letter(2+i)}{int_ded_row}+{get_column_letter(2+i)}{salt_row}")
        c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.border = BOX
    it_row = r
    r += 1

    # Standard deduction
    ws_tx.cell(row=r, column=1, value="Standard deduction (2024)").font = BLACK
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f'=IF({named["filing"]}="M",29200,14600)')
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    std_row = r
    r += 1

    # Tax savings
    ws_tx.cell(row=r, column=1, value="Federal tax savings").font = BLACK_BOLD
    for i, y in enumerate(yrs):
        c = ws_tx.cell(row=r, column=2+i,
                       value=f"=MAX(0,{get_column_letter(2+i)}{it_row}-{get_column_letter(2+i)}{std_row})*{named['marg_rate']}")
        c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX

    r += 2
    ws_tx.cell(row=r, column=1, value="CAPITAL GAINS AT SALE (Year 30)").font = SECTION
    r += 1

    cg_items = [
        ("Sale value (year 30)",      f"=Appreciation!B34", "Home appreciated at the rate in Inputs"),
        ("Selling costs",             f"=Appreciation!E34", "Pct of sale value"),
        ("Cost basis (price + closing)", f"={named['home_price']}+{named['closing']}", "Original price + closing costs"),
        ("Gross gain",                f"=MAX(0,B{r}-B{r+1}-B{r+2})", "Sale − selling costs − basis"),
        ("Section 121 exclusion",     f'=IF({named["filing"]}="M",500000,250000)', "$500K MFJ / $250K single"),
        ("Taxable gain",              f"=MAX(0,B{r+3}-B{r+4})", "Excess over exclusion"),
        ("Federal LTCG tax",          f"=B{r+5}*{named['ltcg_rate']}", "Taxable gain × LTCG rate"),
        ("Net equity (pre-tax)",      f"=Appreciation!F34", "From Appreciation sheet"),
        ("After-tax net equity",      f"=B{r+7}-B{r+6}", "What you actually walk away with"),
    ]
    for i, (label, formula, note) in enumerate(cg_items):
        ws_tx.cell(row=r+i, column=1, value=label).font = BLACK
        c = ws_tx.cell(row=r+i, column=2, value=formula)
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
        ws_tx.cell(row=r+i, column=6, value=note).font = SUBTITLE
    # Highlight after-tax net equity
    ws_tx.cell(row=r+len(cg_items)-1, column=1).font = BLACK_BOLD
    ws_tx.cell(row=r+len(cg_items)-1, column=2).font = BLACK_BOLD
    ws_tx.cell(row=r+len(cg_items)-1, column=2).fill = LIGHT_GREEN

    # ==================== RETIREMENT TRADEOFF SHEET ====================
    ws_rt = wb.create_sheet("Retirement Tradeoff")
    set_widths(ws_rt, [40, 18, 40])
    ws_rt["A1"] = "Retirement Tradeoff Analysis"
    ws_rt["A1"].font = TITLE
    ws_rt["A2"] = "How your housing costs affect retirement savings capacity."
    ws_rt["A2"].font = SUBTITLE

    r = 4
    ws_rt.cell(row=r, column=1, value="CURRENT RETIREMENT POSITION").font = SECTION
    r += 1

    ws_rt.cell(row=r, column=1, value="Gross monthly income").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['income']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['monthly_gross'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="Current 401k contribution ($/mo)").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['income']}*{named['retire_pct']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['monthly_401k'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="Employer match received ($/mo)").font = BLACK
    c = ws_rt.cell(row=r, column=2,
                   value=f"=MIN({named['income']}*{named['retire_pct']},{named['income']}*{named['match_cap']})*{named['emp_match']}/12")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['monthly_match'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="Total going to retirement ($/mo)").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2, value=f"={named['monthly_401k']}+{named['monthly_match']}")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX
    named['total_retire_mo'] = f"'Retirement Tradeoff'!$B${r}"
    r += 2

    ws_rt.cell(row=r, column=1, value="HOUSING COST IMPACT").font = SECTION
    r += 1

    ws_rt.cell(row=r, column=1, value="Monthly housing (PITI)").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['monthly_total']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1

    ws_rt.cell(row=r, column=1, value="Housing as % of gross income").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2, value=f"={named['monthly_total']}/{named['monthly_gross']}")
    c.font = BLACK_BOLD; c.number_format = PCT_FMT; c.border = BOX
    ws_rt.cell(row=r, column=3, value="Target: <28% (ideal), <35% (max)").font = SUBTITLE
    named['housing_pct'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="Retirement as % of gross income").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['total_retire_mo']}/{named['monthly_gross']}")
    c.font = BLACK; c.number_format = PCT_FMT; c.border = BOX
    ws_rt.cell(row=r, column=3, value="Target: 15-20% including match").font = SUBTITLE
    r += 1

    ws_rt.cell(row=r, column=1, value="Combined (housing + retirement %)").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2, value=f"={named['housing_pct']}+{named['retire_pct']}")
    c.font = BLACK_BOLD; c.number_format = PCT_FMT; c.border = BOX
    r += 2

    ws_rt.cell(row=r, column=1, value="OPPORTUNITY COST").font = SECTION
    r += 1

    ws_rt.cell(row=r, column=1, value="Current rent (from inputs)").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['rent_mo']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1

    ws_rt.cell(row=r, column=1, value="Monthly difference (own − rent)").font = BLACK
    c = ws_rt.cell(row=r, column=2, value=f"={named['monthly_total']}-{named['rent_mo']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['own_rent_diff'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="If diff invested monthly for N years").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2,
                   value=f"=FV({named['retire_ret']}/12,{named['retire_yrs']}*12,-{named['own_rent_diff']},0)")
    c.font = BLACK_BOLD; c.fill = LIGHT_BLUE; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_rt.cell(row=r, column=3, value="What you'd have if renting + investing the difference").font = SUBTITLE
    named['fv_diff'] = f"'Retirement Tradeoff'!$B${r}"
    r += 2

    ws_rt.cell(row=r, column=1, value="RETIREMENT BALANCE PROJECTION").font = SECTION
    r += 1

    ws_rt.cell(row=r, column=1, value="Current contribution trajectory").font = BLACK
    c = ws_rt.cell(row=r, column=2,
                   value=f"=FV({named['retire_ret']}/12,{named['retire_yrs']}*12,-{named['total_retire_mo']},0)")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_rt.cell(row=r, column=3, value=f"FV of current contributions over retirement years").font = SUBTITLE
    named['proj_current'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="If you maxed 401k ($23K/yr) + match").font = BLACK
    c = ws_rt.cell(row=r, column=2,
                   value=f"=FV({named['retire_ret']}/12,{named['retire_yrs']}*12,-(23000/12+MIN(23000/12,{named['income']}*{named['match_cap']}/12)*{named['emp_match']}),0)")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['proj_max'] = f"'Retirement Tradeoff'!$B${r}"
    r += 1

    ws_rt.cell(row=r, column=1, value="Retirement shortfall from current plan").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2, value=f"={named['proj_max']}-{named['proj_current']}")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_rt.cell(row=r, column=3, value="Difference between max and current trajectory").font = SUBTITLE
    r += 2

    ws_rt.cell(row=r, column=1, value="VERDICT").font = SECTION
    r += 1
    ws_rt.cell(row=r, column=1, value="Assessment").font = BLACK_BOLD
    c = ws_rt.cell(row=r, column=2,
                   value=f'=IF({named["housing_pct"]}>0.35,"Housing crowding retirement",IF({named["retire_pct"]}<{named["match_cap"]},"Increase 401k to capture full match",IF({named["housing_pct"]}<0.28,"Good balance","Acceptable but watch spending")))')
    c.font = BLACK_BOLD; c.alignment = Alignment(horizontal="center"); c.border = BOX

    # ==================== ROI ANALYSIS SHEET ====================
    ws_roi = wb.create_sheet("ROI Analysis")
    set_widths(ws_roi, [42, 16, 16, 16, 16, 16])
    ws_roi["A1"] = "ROI Analysis — Return on Investment at Sale"
    ws_roi["A1"].font = TITLE
    ws_roi["A2"] = "Three ways to measure your return if you sell the home at different points."
    ws_roi["A2"].font = SUBTITLE

    # Header row
    r = 4
    header_row(ws_roi, r, ["Metric", "Year 5", "Year 10", "Year 15", "Year 20", "Year 30"])
    r += 1

    years = [5, 10, 15, 20, 30]

    # Initial cash outlay (same for all years)
    ws_roi.cell(row=r, column=1, value="Initial cash outlay (DP + closing)").font = BLACK
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i, value=f"={named['down_amt']}+{named['closing']}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['roi_initial'] = f"'ROI Analysis'!$B${r}"
    r += 1

    # Cumulative payments - sum from amortization for each year
    ws_roi.cell(row=r, column=1, value="Cumulative payments (P&I+Tax+Ins+PMI+HOA)").font = BLACK
    for i, yr in enumerate(years):
        # Sum payments from amortization: D (payment) + tax + ins + PMI months * pmi_mo + hoa
        # Simplified: use total monthly × months as approximation since exact calc is complex
        m_end = min(yr * 12, 360)
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"=SUM(Amortization!D5:D{4+m_end})+SUM(Amortization!G5:G{4+m_end})"
                             f"+{named['tax_mo']}*{m_end}+{named['ins_mo']}*{m_end}"
                             f"+SUMPRODUCT(Amortization!H5:H{4+m_end})*{named['pmi_mo']}+{named['hoa_mo_d']}*{m_end}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1

    # Cumulative maintenance
    ws_roi.cell(row=r, column=1, value="Cumulative maintenance").font = BLACK
    for i, yr in enumerate(years):
        # Approximation: average home value over period × maintenance rate × years
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"={named['maint']}*{named['home_price']}*((1+{named['appr']})^{yr}+1)/2*{yr}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1

    # Total cash invested
    ws_roi.cell(row=r, column=1, value="Total cash invested").font = BLACK_BOLD
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i, value=f"=B{r-3}+{get_column_letter(2+i)}{r-2}+{get_column_letter(2+i)}{r-1}")
        c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.border = BOX; c.fill = LIGHT_BLUE
    total_cash_row = r
    r += 1

    # Home value at sale
    ws_roi.cell(row=r, column=1, value="Home value at sale").font = BLACK
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i, value=f"={named['home_price']}*(1+{named['appr']})^{yr}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    home_value_row = r
    r += 1

    # Loan balance at sale
    ws_roi.cell(row=r, column=1, value="Loan balance at sale").font = BLACK
    for i, yr in enumerate(years):
        m_end = min(yr * 12, 360)
        c = ws_roi.cell(row=r, column=2+i, value=f"=IFERROR(INDEX(Amortization!I:I,{4+m_end}),0)")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    balance_row = r
    r += 1

    # Selling costs
    ws_roi.cell(row=r, column=1, value="Selling costs").font = BLACK
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i, value=f"={get_column_letter(2+i)}{home_value_row}*{named['sell_pct']}")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    sell_cost_row = r
    r += 1

    # Net equity at sale
    ws_roi.cell(row=r, column=1, value="Net equity at sale (after selling costs)").font = BLACK_BOLD
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"={get_column_letter(2+i)}{home_value_row}-{get_column_letter(2+i)}{balance_row}-{get_column_letter(2+i)}{sell_cost_row}")
        c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.border = BOX; c.fill = LIGHT_GREEN
    equity_row = r
    r += 2

    # Return Metrics section
    ws_roi.cell(row=r, column=1, value="RETURN METRICS").font = SECTION
    r += 1

    # 1. Simple ROI
    ws_roi.cell(row=r, column=1, value="1. Simple ROI").font = BLACK
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"=({get_column_letter(2+i)}{equity_row}-{get_column_letter(2+i)}{total_cash_row})/{get_column_letter(2+i)}{total_cash_row}")
        c.font = BLACK; c.number_format = PCT_FMT; c.border = BOX
    ws_roi.cell(row=r, column=7, value="(Net Equity − Total Cash) / Total Cash").font = SUBTITLE
    r += 1

    # 2. CAGR on initial cash
    ws_roi.cell(row=r, column=1, value="2. CAGR (on initial cash only)").font = BLACK
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"=IFERROR(({get_column_letter(2+i)}{equity_row}/{named['roi_initial']})^(1/{yr})-1,0)")
        c.font = BLACK; c.number_format = PCT_FMT; c.border = BOX
    ws_roi.cell(row=r, column=7, value="(Net Equity / Initial Cash)^(1/years) − 1").font = SUBTITLE
    r += 1

    # 3. IRR - Excel has XIRR but we'll use a simplified approximation
    # For IRR, we need to use Excel's IRR or XIRR function with cash flows
    # This is complex to set up in a formula, so we'll note it requires a helper range
    ws_roi.cell(row=r, column=1, value="3. IRR (estimated)").font = BLACK_BOLD
    for i, yr in enumerate(years):
        # Approximation: use a simplified IRR estimate
        # IRR ≈ (Net Equity / Total Cash)^(1/years) - 1 adjusted for timing
        # More accurate would require XIRR with actual cash flow dates
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"=IFERROR(RATE({yr}*12,-({get_column_letter(2+i)}{total_cash_row}-{named['roi_initial']})/{yr}/12,{named['roi_initial']},-{get_column_letter(2+i)}{equity_row})*12,0)")
        c.font = BLACK_BOLD; c.number_format = PCT_FMT; c.border = BOX; c.fill = LIGHT_GREEN
    ws_roi.cell(row=r, column=7, value="Annualized return on all cash flows (estimated)").font = SUBTITLE
    r += 2

    # Comparison section
    ws_roi.cell(row=r, column=1, value="MARKET COMPARISON").font = SECTION
    r += 1

    # If invested in market instead
    ws_roi.cell(row=r, column=1, value="If total cash invested in market instead").font = BLACK
    for i, yr in enumerate(years):
        # FV of initial lump sum + monthly contributions
        # Simplified: initial * (1+r)^y + monthly PMT * FV factor
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"={named['roi_initial']}*(1+{named['inv_ret']})^{yr}"
                             f"+FV({named['inv_ret']}/12,{yr}*12,-({get_column_letter(2+i)}{total_cash_row}-{named['roi_initial']})/{yr}/12,0)")
        c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    market_row = r
    r += 1

    # Difference
    ws_roi.cell(row=r, column=1, value="Home equity vs. market investment").font = BLACK_BOLD
    for i, yr in enumerate(years):
        c = ws_roi.cell(row=r, column=2+i,
                       value=f"={get_column_letter(2+i)}{equity_row}-{get_column_letter(2+i)}{market_row}")
        c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.border = BOX; c.fill = LIGHT_GREEN
    ws_roi.cell(row=r, column=7, value="Positive = home wins, Negative = market wins").font = SUBTITLE

    # ==================== REFINANCE SHEET ====================
    ws_rf = wb.create_sheet("Refinance")
    set_widths(ws_rf, [38, 18, 18, 18, 40])
    ws_rf["A1"] = "Refinance Analysis"
    ws_rf["A1"].font = TITLE
    ws_rf["A2"] = "Compares refinancing at the chosen year vs. keeping the original loan."
    ws_rf["A2"].font = SUBTITLE

    # Status banner driven by the enable toggle
    ws_rf["A3"] = f'=IF({named["refi_enabled"]}=1,"Status: refi simulation ENABLED","Status: refi DISABLED — set Include refi? to 1 on Inputs sheet to enable")'
    ws_rf["A3"].font = BLACK_BOLD
    ws_rf["A3"].fill = LIGHT_GREEN

    r = 4
    # Pull balance at refi point from amortization
    ws_rf.cell(row=r, column=1, value="Balance at refi point").font = BLACK
    c = ws_rf.cell(row=r, column=2,
                   value=f"=INDEX(Amortization!I:I,4+{named['refi_year']}*12)")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['refi_bal'] = f"Refinance!$B${r}"
    r += 1

    ws_rf.cell(row=r, column=1, value="Cash-out amount").font = BLACK
    c = ws_rf.cell(row=r, column=2, value=f"={named['cashout']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    r += 1

    ws_rf.cell(row=r, column=1, value="Refi closing costs").font = BLACK
    c = ws_rf.cell(row=r, column=2,
                   value=f"=({named['refi_bal']}+{named['cashout']})*{named['refi_close']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['refi_close_amt'] = f"Refinance!$B${r}"
    r += 1

    ws_rf.cell(row=r, column=1, value="New loan amount").font = BLACK_BOLD
    c = ws_rf.cell(row=r, column=2,
                   value=f"={named['refi_bal']}+{named['cashout']}+IF({named['roll_closing']}=1,{named['refi_close_amt']},0)")
    c.font = BLACK_BOLD; c.number_format = CURRENCY_FMT; c.fill = LIGHT_BLUE; c.border = BOX
    named['new_loan'] = f"Refinance!$B${r}"
    r += 1

    ws_rf.cell(row=r, column=1, value="Cash needed at refi (out of pocket)").font = BLACK
    c = ws_rf.cell(row=r, column=2,
                   value=f"=IF({named['roll_closing']}=1,0,{named['refi_close_amt']})")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    named['refi_oop'] = f"Refinance!$B${r}"
    r += 2

    # Comparison table
    header_row(ws_rf, r, ["Metric", "Keep current loan", "Refinance", "Delta"])
    r += 1

    # Old monthly P&I = current pi
    ws_rf.cell(row=r, column=1, value="Monthly P&I").font = BLACK
    c = ws_rf.cell(row=r, column=2, value=f"={named['pi']}")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    c2 = ws_rf.cell(row=r, column=3, value=f"=-PMT({named['refi_rate']}/12,{named['refi_term']}*12,{named['new_loan']})")
    c2.font = BLACK; c2.number_format = CURRENCY_FMT; c2.border = BOX
    c3 = ws_rf.cell(row=r, column=4, value=f"=C{r}-B{r}")
    c3.font = BLACK; c3.number_format = CURRENCY_FMT; c3.border = BOX
    named['new_pi'] = f"Refinance!$C${r}"
    r += 1

    # Months remaining
    ws_rf.cell(row=r, column=1, value="Months remaining").font = BLACK
    ws_rf.cell(row=r, column=2, value=f"={named['term']}*12-{named['refi_year']}*12").number_format = "0"
    ws_rf.cell(row=r, column=3, value=f"={named['refi_term']}*12").number_format = "0"
    ws_rf.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "0"
    r += 1

    # Total interest remaining (old)
    ws_rf.cell(row=r, column=1, value="Total interest remaining").font = BLACK
    m_start = 4 + 1  # all interest from row 5
    # Old interest remaining = SUM of interest from month refi_year*12+1 to 360
    # Using formula approximation: total loan interest * fraction remaining; OR use SUM with INDEX
    ws_rf.cell(row=r, column=2,
               value=f"=SUM(INDEX(Amortization!F:F,4+{named['refi_year']}*12+1):INDEX(Amortization!F:F,4+{named['term']}*12))")
    ws_rf.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws_rf.cell(row=r, column=3,
               value=f"={named['new_pi']}*{named['refi_term']}*12-{named['new_loan']}")
    ws_rf.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_rf.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = CURRENCY_FMT
    r += 1

    # Total cost remaining (P+I)
    ws_rf.cell(row=r, column=1, value="Total P+I remaining").font = BLACK
    ws_rf.cell(row=r, column=2,
               value=f"={named['pi']}*({named['term']}*12-{named['refi_year']}*12)")
    ws_rf.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws_rf.cell(row=r, column=3,
               value=f"={named['new_pi']}*{named['refi_term']}*12+{named['refi_oop']}")
    ws_rf.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_rf.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = CURRENCY_FMT
    r += 2

    # Summary metrics
    ws_rf.cell(row=r, column=1, value="SUMMARY").font = SECTION
    r += 1
    ws_rf.cell(row=r, column=1, value="Monthly savings").font = BLACK_BOLD
    c = ws_rf.cell(row=r, column=2, value=f"={named['pi']}-{named['new_pi']}")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = CURRENCY_FMT; c.border = BOX
    named['refi_monthly_savings'] = f"Refinance!$B${r}"
    r += 1

    ws_rf.cell(row=r, column=1, value="Break-even (months)").font = BLACK_BOLD
    c = ws_rf.cell(row=r, column=2,
                   value=f"=IFERROR(IF({named['refi_monthly_savings']}<=0,\"Never\",IF({named['refi_oop']}=0,0,{named['refi_oop']}/{named['refi_monthly_savings']})),\"Never\")")
    c.font = BLACK_BOLD; c.fill = LIGHT_GREEN; c.number_format = "0.0"; c.border = BOX
    r += 1

    ws_rf.cell(row=r, column=1, value="Lifetime interest saved").font = BLACK
    c = ws_rf.cell(row=r, column=2,
                   value=f"=SUM(INDEX(Amortization!F:F,4+{named['refi_year']}*12+1):INDEX(Amortization!F:F,4+{named['term']}*12))-({named['new_pi']}*{named['refi_term']}*12-{named['new_loan']})")
    c.font = BLACK; c.number_format = CURRENCY_FMT; c.border = BOX
    ws_rf.cell(row=r, column=5, value="Positive = saves; negative = extending term costs more total interest").font = SUBTITLE
    r += 1

    ws_rf.cell(row=r, column=1, value="Verdict").font = BLACK_BOLD
    c = ws_rf.cell(row=r, column=2,
                   value=f'=IF({named["refi_enabled"]}=0,"Disabled (toggle on Inputs)",IF({named["refi_rate"]}>={named["rate"]},"Rate not lower — only refi if cash-out needed",IF({named["refi_monthly_savings"]}<=0,"Higher monthly cost",IF(B{r-1}<0,"Lower monthly but more total interest","Refi favorable"))))')
    c.font = BLACK_BOLD; c.alignment = Alignment(horizontal="center"); c.border = BOX

    # ==================== EXTRA PAYMENT SHEET ====================
    ws_e = wb.create_sheet("Extra Payment")
    set_widths(ws_e, [38, 18, 18, 30])
    ws_e["A1"] = "Extra Payment Impact"
    ws_e["A1"].font = TITLE
    ws_e["A2"] = "Compares 'Current' (with extra) to 'Baseline' (no extra) using the loan you set in Inputs."
    ws_e["A2"].font = SUBTITLE

    ws_e["A4"] = ""
    ws_e["B4"] = "Baseline (no extra)"
    ws_e["B4"].font = BLACK_BOLD
    ws_e["C4"] = "With extra"
    ws_e["C4"].font = BLACK_BOLD
    ws_e["B4"].fill = HEADER_FILL; ws_e["B4"].font = HEADER
    ws_e["C4"].fill = HEADER_FILL; ws_e["C4"].font = HEADER

    # Baseline monthly P&I + standard loan formulas (NPER, total interest)
    r = 5
    rows_ex = [
        ("Loan amount",            f"={named['loan']}",            f"={named['loan']}"),
        ("Monthly P&I",            f"=-PMT({named['monthly_rate']},{named['n_pmts']},{named['loan']})",
                                   f"=-PMT({named['monthly_rate']},{named['n_pmts']},{named['loan']})"),
        ("Extra per month",        "0",                            f"={named['extra']}"),
        ("Effective payment",      f"=B{r+1}",                     f"=C{r+1}+C{r+2}"),
        # Months to payoff for baseline = full term, for extra use NPER (estimate)
        ("Months to payoff",       f"={named['n_pmts']}",
                                   f"=IFERROR(NPER({named['monthly_rate']},-(B{r+1}+{named['extra']}),{named['loan']}),{named['n_pmts']})"),
        ("Years to payoff",        f"=B{r+4}/12",                  f"=C{r+4}/12"),
        ("Total interest",         f"=B{r+1}*B{r+4}-{named['loan']}",
                                   f"=(B{r+1}+{named['extra']})*C{r+4}-{named['loan']}"),
        ("Total paid",             f"=B{r+1}*B{r+4}",              f"=(B{r+1}+{named['extra']})*C{r+4}"),
    ]
    labels = ["Loan amount", "Monthly P&I", "Extra per month", "Effective payment",
              "Months to payoff", "Years to payoff", "Total interest", "Total paid"]
    formats = [CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT,
               "0.0", "0.0", CURRENCY_FMT, CURRENCY_FMT]

    for i, lbl in enumerate(labels):
        rr = r + i
        ws_e.cell(row=rr, column=1, value=lbl).font = BLACK
        ws_e.cell(row=rr, column=2, value=rows_ex[i][1]).number_format = formats[i]
        ws_e.cell(row=rr, column=3, value=rows_ex[i][2]).number_format = formats[i]
        ws_e.cell(row=rr, column=2).font = BLACK
        ws_e.cell(row=rr, column=3).font = BLACK

    # Savings row
    save_row = r + len(labels) + 1
    ws_e.cell(row=save_row, column=1, value="Interest saved").font = BLACK_BOLD
    c = ws_e.cell(row=save_row, column=2, value=f"=B{r+6}-C{r+6}")
    c.number_format = CURRENCY_FMT; c.font = BLACK_BOLD; c.fill = LIGHT_GREEN

    ws_e.cell(row=save_row+1, column=1, value="Months saved").font = BLACK_BOLD
    c = ws_e.cell(row=save_row+1, column=2, value=f"=B{r+4}-C{r+4}")
    c.number_format = "0.0"; c.font = BLACK_BOLD; c.fill = LIGHT_GREEN

    ws_e.cell(row=save_row+2, column=1, value="Years saved").font = BLACK_BOLD
    c = ws_e.cell(row=save_row+2, column=2, value=f"=B{r+5}-C{r+5}")
    c.number_format = "0.0"; c.font = BLACK_BOLD; c.fill = LIGHT_GREEN

    # ==================== SCENARIOS SHEET ====================
    ws_s = wb.create_sheet("Scenarios")
    set_widths(ws_s, [32, 16, 16, 16])
    ws_s["A1"] = "Scenario Comparison"
    ws_s["A1"].font = TITLE
    ws_s["A2"] = "Three editable scenarios. Edit inputs in B/C/D columns to compare side-by-side."
    ws_s["A2"].font = SUBTITLE

    # Header
    ws_s["A4"] = "Assumption"; ws_s["A4"].font = HEADER; ws_s["A4"].fill = HEADER_FILL
    ws_s["B4"] = "Scenario A"; ws_s["B4"].font = HEADER; ws_s["B4"].fill = HEADER_FILL
    ws_s["C4"] = "Scenario B"; ws_s["C4"].font = HEADER; ws_s["C4"].fill = HEADER_FILL
    ws_s["D4"] = "Scenario C"; ws_s["D4"].font = HEADER; ws_s["D4"].fill = HEADER_FILL

    # Three scenarios — defaults
    scenario_inputs = [
        ("Home price ($)", 500_000, 500_000, 600_000, CURRENCY_FMT),
        ("Down payment (%)", 0.20, 0.10, 0.20, PCT_FMT),
        ("Interest rate", 0.065, 0.065, 0.065, PCT_FMT),
        ("Loan term (years)", 30, 30, 15, "0"),
        ("Extra/month ($)", 0, 0, 500, CURRENCY_FMT),
        ("Appreciation (annual)", 0.035, 0.035, 0.035, PCT_FMT),
    ]
    s_row = 5
    for label, a, b, c, fmt in scenario_inputs:
        ws_s.cell(row=s_row, column=1, value=label).font = BLACK
        for j, val in enumerate([a, b, c]):
            cell = ws_s.cell(row=s_row, column=2+j, value=val)
            cell.font = BLUE; cell.fill = YELLOW_FILL; cell.number_format = fmt
            cell.border = BOX
        s_row += 1

    # Outputs section
    s_row += 1
    ws_s.cell(row=s_row, column=1, value="RESULTS").font = SECTION
    s_row += 1
    outputs_start = s_row

    # Helper indexes — assumes the order in scenario_inputs above (price=5, down%=6, rate=7, term=8, extra=9, appr=10)
    PRICE_R, DOWN_R, RATE_R, TERM_R, EXTRA_R, APPR_R = 5, 6, 7, 8, 9, 10

    outputs = [
        ("Down payment ($)", lambda col: f"={col}{PRICE_R}*{col}{DOWN_R}", CURRENCY_FMT),
        ("Loan amount ($)",  lambda col: f"={col}{PRICE_R}-{col}{PRICE_R}*{col}{DOWN_R}", CURRENCY_FMT),
        ("Monthly P&I",      lambda col: f"=-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))", CURRENCY_FMT),
        ("Monthly tax+ins",  lambda col: f"={named['tax_mo']}+{named['ins_mo']}", CURRENCY_FMT),
        ("Total monthly (est)", lambda col: f"=-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}", CURRENCY_FMT),
        ("Total interest (no extra)", lambda col: f"=-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))*{col}{TERM_R}*12-({col}{PRICE_R}*(1-{col}{DOWN_R}))", CURRENCY_FMT),
        ("Months to payoff (w/ extra)",
            lambda col: f"=IFERROR(NPER({col}{RATE_R}/12,-(-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{col}{EXTRA_R}),{col}{PRICE_R}*(1-{col}{DOWN_R})),{col}{TERM_R}*12)", "0.0"),
        ("Years to payoff (w/ extra)",
            lambda col: f"=IFERROR(NPER({col}{RATE_R}/12,-(-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{col}{EXTRA_R}),{col}{PRICE_R}*(1-{col}{DOWN_R}))/12,{col}{TERM_R})", "0.0"),
        ("Home value @ year 30", lambda col: f"={col}{PRICE_R}*(1+{col}{APPR_R})^30", CURRENCY_FMT),
        ("Home value @ year 10", lambda col: f"={col}{PRICE_R}*(1+{col}{APPR_R})^10", CURRENCY_FMT),
        ("Min income (Ramsey 25%)",
            lambda col: f"=MAX((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))/0.25,((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))+{named['debts_mo']})/0.25)*12",
            CURRENCY_FMT),
        ("Min income (28/36 rule)",
            lambda col: f"=MAX((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))/0.28,((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))+{named['debts_mo']})/0.36)*12",
            CURRENCY_FMT),
        ("Min income (FHA 31/43)",
            lambda col: f"=MAX((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))/0.31,((-PMT({col}{RATE_R}/12,{col}{TERM_R}*12,{col}{PRICE_R}*(1-{col}{DOWN_R}))+{named['tax_pct']}*{col}{PRICE_R}/12+{named['ins_mo']}+{named['hoa_mo_d']}+IF({col}{DOWN_R}<0.2,{named['pmi_pct']}*{col}{PRICE_R}*(1-{col}{DOWN_R})/12,0))+{named['debts_mo']})/0.43)*12",
            CURRENCY_FMT),
    ]
    for label, formula_fn, fmt in outputs:
        ws_s.cell(row=s_row, column=1, value=label).font = BLACK
        for j, col in enumerate(["B", "C", "D"]):
            c = ws_s.cell(row=s_row, column=2+j, value=formula_fn(col))
            c.font = BLACK
            c.number_format = fmt
            c.border = BOX
        s_row += 1

    # Highlight key row
    for col in range(1, 5):
        ws_s.cell(row=outputs_start+4, column=col).fill = LIGHT_GREEN  # Total monthly
        ws_s.cell(row=outputs_start+8, column=col).fill = LIGHT_BLUE  # Home value yr 30

    # ==================== README SHEET ====================
    ws_h = wb.create_sheet("README", 0)  # Make first sheet
    set_widths(ws_h, [100])
    ws_h["A1"] = "Home Mortgage Calculator"
    ws_h["A1"].font = TITLE
    notes = [
        "",
        "How to use:",
        "  1. Open the Inputs sheet and edit the YELLOW cells (blue text).",
        "  2. All other sheets update automatically.",
        "",
        "Sheets:",
        "  • Inputs           — All editable assumptions. Includes monthly payment summary.",
        "  • Amortization     — Full 360-month schedule. PMI auto-deactivates at 20% LTV.",
        "  • Appreciation     — Year-by-year asset value, equity, and cumulative interest. Includes chart.",
        "  • Rent vs Buy      — Net wealth path (with escalating tax + insurance). Includes chart.",
        "  • Affordability    — Minimum household income required at conventional DTI standards.",
        "  • Cash to Close    — Cash needed at closing and recommended liquid retention.",
        "  • Buy Today        — Liquidity check and concentration cap analysis.",
        "  • Tax Benefits     — Mortgage interest + SALT deductions, cap gains exclusion at sale.",
        "  • Retirement Tradeoff — How housing costs affect retirement savings capacity.",
        "  • ROI Analysis     — Return on investment (Simple ROI, CAGR, IRR) at years 5/10/15/20/30.",
        "  • Refinance        — Compare refinancing at year N vs. keeping current loan.",
        "  • Extra Payment    — How much interest you save by paying extra each month.",
        "  • Scenarios        — Three editable side-by-side scenarios for what-if analysis.",
        "",
        "Color conventions:",
        "  • Blue text on yellow = editable input",
        "  • Black text         = formula",
        "  • Green text         = cross-sheet reference",
        "",
        "Assumptions baked into the model:",
        "  • Fixed-rate mortgage (single rate for whole term)",
        "  • PMI applied while equity < 20% of original purchase price (uses standard cutoff)",
        "  • Property tax % is applied to original home price (constant). For tax-on-current-value, edit Inputs!B17.",
        "  • Selling costs (6% default) deducted only when computing net equity at exit.",
        "  • Rent vs. buy: renter invests the down payment + closing costs upfront at the investment return rate. Differential method NOT used — a stricter model would also invest the difference between rent and total monthly carrying cost.",
        "",
        "Caveats and ideas for v2:",
        "  • Tax deductibility of mortgage interest (post-TCJA SALT cap)",
        "  • ARM / rate change after fixed period",
        "  • Refi scenarios",
        "  • Capital gains exclusion at sale ($250k / $500k)",
        "  • Cost basis adjustments for capital improvements",
        "",
    ]
    for i, line in enumerate(notes):
        ws_h.cell(row=2+i, column=1, value=line)
        if line.endswith(":") and not line.startswith(" "):
            ws_h.cell(row=2+i, column=1).font = SECTION

    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(script_dir, "mortgage_calculator.xlsx")
    wb.save(out_path)
    print(f"Saved to {out_path}")


if __name__ == "__main__":
    build()
